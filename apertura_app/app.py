"""
Apertura de Cuentas INVEX — Persona Física y Persona Moral
App Streamlit: sube el PDF → extrae → llena formatos → descarga ZIP
"""
from __future__ import annotations

import io
import re
import abc
import shutil
import zipfile
import tempfile
from dataclasses import dataclass, field
from datetime import date, datetime
from enum import Enum
from pathlib import Path
from typing import Optional

import streamlit as st

# ── página ────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Apertura de Cuentas INVEX",
    page_icon="🏦",
    layout="centered",
)

# ── estilos ───────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Serif+Display&family=DM+Sans:wght@300;400;500;600&display=swap');

html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }

.block-container { max-width: 780px; padding-top: 2.5rem; }

h1 { font-family: 'DM Serif Display', serif; font-size: 2.2rem !important;
     color: #0a2540 !important; letter-spacing: -0.5px; }

h3 { font-family: 'DM Serif Display', serif; color: #0a2540 !important; }

.stButton > button {
    background: #0a2540; color: #fff; border: none;
    border-radius: 8px; padding: 0.6rem 1.6rem;
    font-family: 'DM Sans', sans-serif; font-weight: 500;
    font-size: 0.95rem; cursor: pointer; transition: background 0.2s;
}
.stButton > button:hover { background: #1a4070; }

.upload-hint {
    background: #f0f4f8; border: 1.5px dashed #c5d2e0;
    border-radius: 12px; padding: 1.5rem;
    text-align: center; color: #4a5568; font-size: 0.9rem;
    margin-bottom: 1rem;
}

.campo-card {
    background: #f8fafc; border: 1px solid #e2e8f0;
    border-radius: 10px; padding: 1rem 1.2rem;
    margin-bottom: 0.5rem;
}
.campo-label { font-size: 0.75rem; color: #718096; font-weight: 500;
               text-transform: uppercase; letter-spacing: 0.05em; }
.campo-valor { font-size: 1rem; color: #0a2540; font-weight: 500; margin-top: 2px; }

.warn-box {
    background: #fffbeb; border: 1px solid #f6d860;
    border-radius: 8px; padding: 0.8rem 1rem;
    color: #7d5b00; font-size: 0.88rem; margin-top: 0.5rem;
}
.ok-box {
    background: #f0fff4; border: 1px solid #68d391;
    border-radius: 8px; padding: 0.8rem 1rem;
    color: #276749; font-size: 0.88rem; margin-top: 0.5rem;
}
.divider { border: none; border-top: 1px solid #e2e8f0; margin: 1.5rem 0; }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# DOMINIO
# ══════════════════════════════════════════════════════════════════════════════

class TipoPersona(str, Enum):
    FISICA = "FISICA"
    MORAL  = "MORAL"

class Sexo(str, Enum):
    MASCULINO = "MASCULINO"
    FEMENINO  = "FEMENINO"

RFC_PF = re.compile(r"^[A-ZÑ&]{4}\d{6}[A-Z0-9]{3}$")
RFC_PM = re.compile(r"^[A-ZÑ&]{3}\d{6}[A-Z0-9]{3}$")
CURP_RE = re.compile(r"^[A-Z]{4}\d{6}[HM][A-Z]{5}[A-Z0-9]\d$")


class Domicilio:
    def __init__(self, calle_numero="", colonia="", municipio="", ciudad="",
                 entidad_federativa="", pais="MEXICO", codigo_postal=""):
        self.calle_numero       = calle_numero
        self.colonia            = colonia
        self.municipio          = municipio
        self.ciudad             = ciudad
        self.entidad_federativa = entidad_federativa
        self.pais               = pais
        self.codigo_postal      = codigo_postal

    def una_linea(self) -> str:
        partes = [self.calle_numero, self.colonia, self.municipio,
                  self.entidad_federativa, self.pais]
        base = ", ".join(p for p in partes if p)
        if self.codigo_postal:
            base += f"  C.P. {self.codigo_postal}"
        return base


class Persona:
    def __init__(self, nombre="", apellido_paterno="", apellido_materno="",
                 rfc="", curp="", fecha_nacimiento=None, sexo=None,
                 nacionalidad="MEXICO", tipo_identificacion="INE",
                 numero_identificacion="", vigencia_identificacion=None,
                 actividad="", fuente_ingresos="", es_pep=False,
                 telefono="", correo="", domicilio=None):
        self.nombre                 = nombre
        self.apellido_paterno       = apellido_paterno
        self.apellido_materno       = apellido_materno
        self.rfc                    = rfc.strip().upper()
        self.curp                   = curp.strip().upper()
        self.fecha_nacimiento       = fecha_nacimiento
        self.sexo                   = sexo
        self.nacionalidad           = nacionalidad.strip().upper()
        self.tipo_identificacion    = tipo_identificacion.strip().upper()
        self.numero_identificacion  = numero_identificacion
        self.vigencia_identificacion= vigencia_identificacion
        self.actividad              = actividad
        self.fuente_ingresos        = fuente_ingresos
        self.es_pep                 = es_pep
        self.telefono               = telefono
        self.correo                 = correo
        self.domicilio              = domicilio or Domicilio()

    @property
    def nombre_completo(self) -> str:
        return " ".join(p for p in [self.nombre, self.apellido_paterno,
                                    self.apellido_materno] if p).strip()


class Empresa:
    def __init__(self, razon_social="", rfc="", nacionalidad="MEXICO",
                 pais_residencia="MEXICO", fecha_constitucion=None,
                 acta_constitutiva="", notario="", numero_notario="",
                 ciudad_notario="", inscripcion_rpp="", tipo_sociedad="",
                 giro="", fiel="", telefono="", correo="", domicilio=None):
        self.razon_social       = razon_social
        self.rfc                = rfc.strip().upper()
        self.nacionalidad       = nacionalidad
        self.pais_residencia    = pais_residencia
        self.fecha_constitucion = fecha_constitucion
        self.acta_constitutiva  = acta_constitutiva
        self.notario            = notario
        self.numero_notario     = numero_notario
        self.ciudad_notario     = ciudad_notario
        self.inscripcion_rpp    = inscripcion_rpp
        self.tipo_sociedad      = tipo_sociedad
        self.giro               = giro
        self.fiel               = fiel
        self.telefono           = telefono
        self.correo             = correo
        self.domicilio          = domicilio or Domicilio()


class Accionista:
    def __init__(self, nombre="", apellido_paterno="", apellido_materno="",
                 porcentaje=0.0, nacionalidad="", es_pep=False):
        self.nombre           = nombre
        self.apellido_paterno = apellido_paterno
        self.apellido_materno = apellido_materno
        self.porcentaje       = porcentaje
        self.nacionalidad     = nacionalidad
        self.es_pep           = es_pep

    @property
    def nombre_completo(self) -> str:
        return " ".join(p for p in [self.nombre, self.apellido_paterno,
                                    self.apellido_materno] if p).strip()


class RepresentanteLegal:
    def __init__(self, nombre="", apellido_paterno="", apellido_materno="",
                 nacionalidad="MEXICO", tipo_identificacion="INE",
                 numero_identificacion="", vigencia_identificacion=None,
                 puesto="", acta_poderes="", notario_poderes="", ciudad_poderes=""):
        self.nombre                  = nombre
        self.apellido_paterno        = apellido_paterno
        self.apellido_materno        = apellido_materno
        self.nacionalidad            = nacionalidad
        self.tipo_identificacion     = tipo_identificacion
        self.numero_identificacion   = numero_identificacion
        self.vigencia_identificacion = vigencia_identificacion
        self.puesto                  = puesto
        self.acta_poderes            = acta_poderes
        self.notario_poderes         = notario_poderes
        self.ciudad_poderes          = ciudad_poderes

    @property
    def nombre_completo(self) -> str:
        return " ".join(p for p in [self.nombre, self.apellido_paterno,
                                    self.apellido_materno] if p).strip()


class Solicitud:
    def __init__(self, tipo_persona, institucion="INVEX", numero_contrato="",
                 fecha_alta=None, ejecutivo_cuenta="", titular=None,
                 cotitulares=None, procedencia_recursos="", uso_cuenta="",
                 monto_apertura=None, empresa=None, accionistas=None,
                 representantes=None):
        self.tipo_persona        = tipo_persona
        self.institucion         = institucion
        self.numero_contrato     = (numero_contrato or "").strip()
        self.fecha_alta          = fecha_alta
        self.ejecutivo_cuenta    = ejecutivo_cuenta
        self.titular             = titular or Persona()
        self.cotitulares         = cotitulares or []
        self.procedencia_recursos= procedencia_recursos
        self.uso_cuenta          = uso_cuenta
        self.monto_apertura      = monto_apertura
        self.empresa             = empresa
        self.accionistas         = accionistas or []
        self.representantes      = representantes or []

    def advertencias(self) -> list[str]:
        w = []
        if not self.numero_contrato:
            w.append("Falta el número de contrato.")
        if self.tipo_persona == TipoPersona.FISICA:
            t = self.titular
            if t.rfc and not (RFC_PF.match(t.rfc) or RFC_PM.match(t.rfc)):
                w.append(f"RFC '{t.rfc}' no cumple el formato oficial.")
            if t.curp and not CURP_RE.match(t.curp):
                w.append(f"CURP '{t.curp}' no cumple el formato oficial.")
            if not t.nombre_completo:
                w.append("Falta el nombre del titular.")
            if t.correo and "@" not in t.correo:
                w.append(f"Correo '{t.correo}' parece inválido.")
        else:
            e = self.empresa
            if not e or not e.razon_social:
                w.append("Falta la razón social de la empresa.")
            if e and e.rfc and not RFC_PM.match(e.rfc):
                w.append(f"RFC '{e.rfc}' no cumple el formato de Persona Moral.")
            if not self.representantes:
                w.append("No se encontró representante legal.")
        return w


# ══════════════════════════════════════════════════════════════════════════════
# EXTRACTOR
# ══════════════════════════════════════════════════════════════════════════════

def _fecha(s):
    if not s: return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try: return datetime.strptime(s.strip(), fmt).date()
        except ValueError: continue
    return None

def _buscar(patron, texto, grupo=1):
    m = re.search(patron, texto, re.IGNORECASE)
    return m.group(grupo).strip() if m else ""

_STOP = (r"TIPO FIRMA|FECHA NACIMIENTO|C\.U\.R\.P|ENT\. NAC|R\. ?F\.C|"
         r"TIPO IDENTIFICACION|NUMERO IDENTIFICACION|SEXO|VIGENCIA|FIEL|TIN|"
         r"CODIGO POSTAL|CD\. O POB|COMPROBANTE|PAIS|NUMERO|RESIDENCIA TELEFONO|"
         r"MEDIO NOTIFICA|EN CASO AFIRMATIVO|\n")


def leer_texto_pdf(data: bytes) -> str:
    import pdfplumber
    paginas = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page in pdf.pages:
            paginas.append(page.extract_text() or "")
    return "\n".join(paginas)


def extraer_pf(data: bytes) -> Solicitud:
    texto = leer_texto_pdf(data)
    titular_txt = texto.split("INFORMACION DE COTITULARES")[0]

    def _persona(t: str) -> Persona:
        sexo_raw = _buscar(r"SEXO:\s*(MASCULINO|FEMENINO)", t)
        dom = Domicilio(
            calle_numero      = _buscar(r"CALLE Y NUMERO:\s*(.+)", t),
            colonia           = _buscar(r"COLONIA:\s*(.+?)(?:" + _STOP + ")", t),
            municipio         = _buscar(r"DELEGACION O MUNICIPIO:\s*(.+?)(?:" + _STOP + ")", t),
            ciudad            = _buscar(r"CD\. O POB\.?:\s*(.+?)(?:" + _STOP + ")", t),
            entidad_federativa= _buscar(r"ENTIDAD FEDERATIVA:\s*(.+?)(?:" + _STOP + ")", t),
            pais              = _buscar(r"PAIS:\s*([A-ZÁÉÍÓÚÑ]+)", t) or "MEXICO",
            codigo_postal     = _buscar(r"CODIGO POSTAL:\s*(\d{4,5})", t),
        )
        tipo_id = _buscar(r"TIPO IDENTIFICACION:\s*(.+?)(?:" + _STOP + ")", t)
        if "ELECTOR" in tipo_id.upper() or "INE" in tipo_id.upper():
            tipo_id = "INE"
        return Persona(
            nombre               = _buscar(r"NOMBRE\(S\):\s*(.+?)(?:" + _STOP + ")", t),
            apellido_paterno     = _buscar(r"APELLIDO PATERNO:\s*(.+?)(?:" + _STOP + ")", t),
            apellido_materno     = _buscar(r"APELLIDO MATERNO:\s*(.+?)(?:AP\.?\s*CASADA|" + _STOP + ")", t),
            rfc                  = _buscar(r"R\. ?F\.C\.?:\s*([A-ZÑ&]{3,4}\d{6}[A-Z0-9]{3})", t),
            curp                 = _buscar(r"C\.U\.R\.P\.?:\s*([A-Z0-9]{18})", t),
            fecha_nacimiento     = _fecha(_buscar(r"FECHA NACIMIENTO:\s*(\d{2}/\d{2}/\d{4})", t)),
            sexo                 = Sexo(sexo_raw) if sexo_raw else None,
            nacionalidad         = _buscar(r"NACIONALIDAD:\s*([A-ZÁÉÍÓÚÑ]+)", t) or "MEXICO",
            tipo_identificacion  = tipo_id or "INE",
            numero_identificacion= _buscar(r"NUMERO IDENTIFICACION:\s*([\dA-Z]+)", t),
            vigencia_identificacion= _fecha(_buscar(r"VIGENCIA:\s*(\d{2}/\d{2}/\d{4})", t)),
            actividad            = _buscar(r"ACTIVIDAD:\s*(.+?)(?:" + _STOP + ")", t),
            fuente_ingresos      = _buscar(r"FUENTE PRINCIPAL DE INGRESOS:\s*(.+?)(?:" + _STOP + ")", t),
            es_pep               = bool(re.search(r"POLITICAMENTE EXPUEST[OA].*?\bSI\b(?!\s*NO)", t)),
            telefono             = _buscar(r"(?:TELEFONO|CELULAR / MOVIL):\s*(\(?\d[\d()\- ]{7,})", t),
            correo               = _buscar(r"CORREO ELECTRONICO:\s*([\w.\-]+@[\w.\-]+)", t),
            domicilio            = dom,
        )

    sol = Solicitud(
        tipo_persona        = TipoPersona.FISICA,
        institucion         = "INVEX",
        numero_contrato     = _buscar(r"\b(\d{8})\b", titular_txt),
        fecha_alta          = _fecha(_buscar(r"\b(\d{2}/\d{2}/\d{4})\b", titular_txt)),
        ejecutivo_cuenta    = _buscar(r"ai\d+\s+([A-ZÁÉÍÓÚÑ ]+?)\s*\n", titular_txt),
        titular             = _persona(titular_txt),
        procedencia_recursos= _buscar(r"PROCEDENCIA DE LOS RECURSOS[^:]*:\s*([A-ZÁÉÍÓÚÑ ]+)", texto),
        uso_cuenta          = _buscar(r"USO QUE SE PRETENDE DAR A LA CUENTA:\s*([A-ZÁÉÍÓÚÑ ]+)", texto),
    )
    if "INFORMACION DE COTITULARES" in texto:
        cot_txt = texto.split("INFORMACION DE COTITULARES")[1].split("CORREOS")[0]
        cot = _persona(cot_txt)
        if cot.nombre_completo:
            sol.cotitulares.append(cot)
    return sol


def extraer_pm(data: bytes) -> Solicitud:
    texto = leer_texto_pdf(data)

    empresa = Empresa(
        razon_social      = _buscar(r"DENOMINACION / RAZON SOCIAL:\s*(.+)", texto),
        rfc               = _buscar(r"R\. ?F\.C\.?:\s*([A-ZÑ&]{3}\d{6}[A-Z0-9]{3})", texto),
        nacionalidad      = _buscar(r"NACIONALIDAD:\s*([A-ZÁÉÍÓÚÑ]+)", texto) or "MEXICO",
        pais_residencia   = _buscar(r"PAIS RESIDENCIA:\s*([A-ZÁÉÍÓÚÑ]+)", texto) or "MEXICO",
        fecha_constitucion= _fecha(_buscar(r"DE FECHA:\s*(\d{2}/\d{2}/\d{4})", texto)),
        acta_constitutiva = _buscar(r"ACTA CONSTITUTUVA No\.:\s*([\d,]+)", texto),
        notario           = _buscar(r"NOTARIO:\s*(.+?)(?:No\.|DE LA CIUDAD|\n)", texto),
        numero_notario    = _buscar(r"No\. NOTARIO:\s*(\d+)", texto),
        ciudad_notario    = _buscar(r"DE LA CIUDAD:\s*([A-ZÁÉÍÓÚÑ]+)", texto),
        inscripcion_rpp   = _buscar(r"INSCRIPCION R\.P\.P\. y C\.:\s*(\d+)", texto),
        tipo_sociedad     = _buscar(r"TIPO DE PERSONA MORAL:\s*(.+?)(?:\n|No\.)", texto),
        giro              = _buscar(r"GIRO DE NEGOCIO / ACTIVIDAD:\s*(.+?)(?:\n|No\.)", texto),
        fiel              = _buscar(r"FIEL:\s*([\dA-Z]+)", texto),
        telefono          = _buscar(r"CELULAR / MOVIL\s*([\d()]+)", texto),
        correo            = _buscar(r"CORREO ELECTRONICO:\s*([\w.\-]+@[\w.\-]+)", texto),
        domicilio         = Domicilio(
            calle_numero      = _buscar(r"CALLE Y NUMERO:\s*(.+)", texto),
            colonia           = _buscar(r"COLONIA:\s*(.+?)(?:CODIGO POSTAL|\n)", texto),
            municipio         = _buscar(r"DELEGACION O MUNICIPIO:\s*(.+?)(?:CD\.|\n)", texto),
            ciudad            = _buscar(r"CD\. O POB\.:\s*(.+?)(?:COMPROBANTE|\n)", texto),
            entidad_federativa= _buscar(r"ENTIDAD FEDERATIVA:\s*(.+?)(?:PAIS|\n)", texto),
            pais              = _buscar(r"PAIS:\s*([A-ZÁÉÍÓÚÑ]+)", texto) or "MEXICO",
            codigo_postal     = _buscar(r"CODIGO POSTAL:\s*(\d{4,5})", texto),
        ),
    )

    # Accionistas
    accionistas = []
    for m in re.finditer(
        r"([A-ZÁÉÍÓÚÑ]+)\s+([A-ZÁÉÍÓÚÑ]+)\s+([A-ZÁÉÍÓÚÑ]+)\s+(\d+)\s+Si\s+No",
        texto, re.IGNORECASE
    ):
        accionistas.append(Accionista(
            nombre           = m.group(1),
            apellido_paterno = m.group(2),
            apellido_materno = m.group(3),
            porcentaje       = float(m.group(4)),
        ))

    # Representante legal
    representantes = []
    if "REPRESENTANTE LEGAL 1" in texto:
        rl_txt = texto.split("REPRESENTANTE LEGAL 1")[1].split("ORGANIGRAMA")[0]
        rl = RepresentanteLegal(
            nombre                = _buscar(r"NOMBRES\(S\):\s*([A-ZÁÉÍÓÚÑ ]+?)(?:NACIONALIDAD|\n)", rl_txt),
            apellido_paterno      = _buscar(r"APELLIDO PATERNO:\s*(.+?)(?:TIPO DE FIRMA|\n)", rl_txt),
            apellido_materno      = _buscar(r"APELLIDO MATERNO:\s*(.+?)(?:PUESTO|\n)", rl_txt),
            nacionalidad          = _buscar(r"NACIONALIDAD:\s*([A-ZÁÉÍÓÚÑ]+)", rl_txt) or "MEXICO",
            tipo_identificacion   = "INE",
            numero_identificacion = _buscar(r"NUMERO IDENTIFICACION:\s*([\dA-Z]+)", rl_txt),
            vigencia_identificacion= _fecha(_buscar(r"VIGENCIA:\s*(\d{2}/\d{2}/\d{4})", rl_txt)),
            puesto                = _buscar(r"PUESTO EN LA EMPRESA:\s*(.+?)(?:\n|$)", rl_txt),
            acta_poderes          = _buscar(r"ACTA No\. \(Poderes\):\s*([\d,]+)", rl_txt),
            notario_poderes       = _buscar(r"ANTE EL NOTARIO:\s*(.+?)(?:CON No\.|\n)", rl_txt),
            ciudad_poderes        = _buscar(r"DE LA CIUDAD:\s*([A-ZÁÉÍÓÚÑ]+)", rl_txt),
        )
        if rl.nombre_completo:
            representantes.append(rl)

    return Solicitud(
        tipo_persona        = TipoPersona.MORAL,
        institucion         = "INVEX",
        numero_contrato     = _buscar(r"\b(\d{8})\b", texto),
        fecha_alta          = _fecha(_buscar(r"\b(\d{2}/\d{2}/\d{4})\b", texto)),
        ejecutivo_cuenta    = _buscar(r"ai\d+\s+([A-ZÁÉÍÓÚÑ ]+?)\s*\n", texto),
        empresa             = empresa,
        accionistas         = accionistas,
        representantes      = representantes,
        procedencia_recursos= _buscar(r"PROCEDENCIA DE LOS RECURSOS[^:]*:\s*([A-ZÁÉÍÓÚÑ ]+)", texto),
        uso_cuenta          = _buscar(r"USO QUE SE PRETENDE DAR A LA CUENTA:\s*([A-ZÁÉÍÓÚÑ ]+)", texto),
    )


def detectar_y_extraer(data: bytes) -> Solicitud:
    texto = leer_texto_pdf(data).upper()
    if "PERSONA MORAL" in texto:
        return extraer_pm(data)
    return extraer_pf(data)



def _f(d): return d.strftime("%d/%m/%Y") if d else ""

def contexto_docx(sol: Solicitud) -> dict:
    t = sol.titular
    nombre = (t.nombre_completo + " / " + sol.cotitulares[0].nombre_completo
              if sol.cotitulares else t.nombre_completo)
    return {
        "contrato": sol.numero_contrato,
        "nombre": nombre,
        "rfc": t.rfc,
        "curp": t.curp,
        "fecha_nacimiento": _f(t.fecha_nacimiento),
        "ine": t.numero_identificacion,
        "actividad": t.actividad,
        "entidad_federativa": t.domicilio.entidad_federativa,
        "pais": t.domicilio.pais,
        "telefono": t.telefono,
        "correo": t.correo,
        "tipo_persona": sol.tipo_persona.value,
        "nacionalidad": t.nacionalidad,
        "institucion": sol.institucion,
        "fecha_alta": _f(sol.fecha_alta),
        "domicilio": t.domicilio.una_linea(),
        "cotitular": (sol.cotitulares[0].nombre_completo if sol.cotitulares else ""),
        "fecha_recepcion": _f(sol.fecha_alta),
    }

def _checklist(sol):
    t = sol.titular
    cot = sol.cotitulares[0].nombre_completo if sol.cotitulares else ""
    return {"Persona Física": {
        "G3": lambda s: s.numero_contrato,
        "G4": lambda s: _f(s.fecha_alta),
        "C10": lambda s: t.nombre_completo,
        "C11": lambda s: cot,
        "C13": lambda s: t.nacionalidad.replace("MEXICO", "MEXICANA"),
        "C14": lambda s: t.domicilio.entidad_federativa,
        "C15": lambda s: ("MUJER" if (t.sexo and t.sexo.value == "FEMENINO") else "HOMBRE"),
        "C16": lambda s: t.actividad,
        "C17": lambda s: t.rfc,
        "C18": lambda s: t.curp,
        "C19": lambda s: t.tipo_identificacion,
        "C20": lambda s: t.numero_identificacion,
        "C24": lambda s: t.domicilio.una_linea(),
        "C26": lambda s: t.telefono,
        "C27": lambda s: t.correo,
    }}

def _kyc(sol):
    t = sol.titular
    return {"Hoja1": {
        "B12": lambda s: t.nombre_completo,
        "D14": lambda s: "FISICA",
        "F15": lambda s: s.monto_apertura or "",
        "E21": lambda s: t.actividad,
        "H33": lambda s: t.correo,
        "C33": lambda s: t.telefono,
        "K12": lambda s: "JUAN JAVIER GILBERTO TELLEZ LOPEZ",
        "B51": lambda s: t.domicilio.calle_numero,
        "F51": lambda s: t.domicilio.colonia,
        "M51": lambda s: t.domicilio.municipio,
        "B53": lambda s: t.domicilio.municipio,
        "E53": lambda s: t.domicilio.entidad_federativa,
        "I53": lambda s: t.domicilio.codigo_postal,
        "K53": lambda s: t.domicilio.pais,        
        "D70": lambda s: t.nombre_completo,
        "D71": lambda s: t.tipo_identificacion,
        "N71": lambda s: t.numero_identificacion,
        "C79": lambda s: t.nombre_completo,
    }}

def _perfil(sol):
    t = sol.titular
    nombre_firma = (" / ".join([t.nombre_completo] +
                   [c.nombre_completo for c in sol.cotitulares])
                   if sol.cotitulares else t.nombre_completo)
    return {"Hoja1": {
        "D146": lambda s: nombre_firma,
        "D147": lambda s: s.numero_contrato,
        "D148": lambda s: sol.fecha_alta,
        "D149": lambda s: t.correo,
        "D158": lambda s: nombre_firma,
    }}

TIPOS_SOCIETARIOS = {
    "S.A. DE C.V.": "SOCIEDAD ANÓNIMA DE CAPITAL VARIABLE",
    "S.A.":         "SOCIEDAD ANÓNIMA",
    "S. DE R.L. DE C.V.": "SOCIEDAD DE RESPONSABILIDAD LIMITADA DE CAPITAL VARIABLE",
    "S. DE R.L.":   "SOCIEDAD DE RESPONSABILIDAD LIMITADA",
    "S.A.S.":       "SOCIEDAD POR ACCIONES SIMPLIFICADA",
    "A.C.":         "ASOCIACIÓN CIVIL",
    "S.C.":         "SOCIEDAD CIVIL",
    "S.A.P.I. DE C.V.": "SOCIEDAD ANÓNIMA PROMOTORA DE INVERSIÓN DE CAPITAL VARIABLE",
}

def _tipo_societario(razon_social: str) -> str:
    """Detecta el tipo societario en la razón social y lo devuelve desarrollado."""
    rs = razon_social.upper()
    for sigla, nombre in TIPOS_SOCIETARIOS.items():
        if sigla.upper() in rs:
            return nombre
    return ""

def contexto_docx_pm(sol: Solicitud) -> dict:
    e = sol.empresa or Empresa()
    rl = sol.representantes[0] if sol.representantes else RepresentanteLegal()
    ac1 = sol.accionistas[0] if len(sol.accionistas) > 0 else Accionista()
    ac2 = sol.accionistas[1] if len(sol.accionistas) > 1 else Accionista()
    # Ciudad del notario: usar ciudad_notario si existe, si no usar municipio del domicilio
    ciudad_n = e.ciudad_notario or e.domicilio.municipio or e.domicilio.ciudad
    entidad  = e.domicilio.entidad_federativa
    return {
        "contrato"          : sol.numero_contrato,
        "razon_social"      : e.razon_social,
        "rfc"               : e.rfc,
        "nacionalidad"      : e.nacionalidad,
        "fecha_constitucion": _f(e.fecha_constitucion),
        "acta_constitutiva" : e.acta_constitutiva,
        "notario"           : e.notario,
        "numero_notario"    : e.numero_notario,
        "ciudad_notario"    : ciudad_n,
        "inscripcion_rpp"   : e.inscripcion_rpp,
        "tipo_societario"   : _tipo_societario(e.razon_social),
        "giro"              : e.giro,
        "fiel"              : e.fiel,
        "telefono"          : e.telefono,
        "correo"            : e.correo,
        "domicilio"         : e.domicilio.una_linea(),
        "pais"              : e.domicilio.pais,
        "entidad_federativa": entidad,
        "fecha_alta"        : _f(sol.fecha_alta),
        "fecha_recepcion"   : _f(sol.fecha_alta),
        "rep_nombre"        : rl.nombre_completo,
        "rep_puesto"        : rl.puesto,
        "rep_ine"           : rl.numero_identificacion,
        "rep_vigencia"      : _f(rl.vigencia_identificacion),
        "acta_poderes"      : rl.acta_poderes,
        "notario_poderes"   : rl.notario_poderes,
        "accionista1"       : ac1.nombre_completo,
        "pct1"              : f"{ac1.porcentaje:.0f}",
        "accionista2"       : ac2.nombre_completo,
        "pct2"              : f"{ac2.porcentaje:.0f}",
        # campos PF vacíos para que Jinja no falle
        "nombre": e.razon_social, "curp": "", "ine": "",
        "fecha_nacimiento": "", "actividad": e.giro,
        "tipo_persona": "MORAL", "institucion": sol.institucion,
        "cotitular": "",
    }

def _checklist_pm(sol):
    e = sol.empresa or Empresa()
    rl = sol.representantes[0] if sol.representantes else RepresentanteLegal()
    LLENAR_POR_ASESOR = "llenarse por SM/ASESOR"
    return {"Persona Moral": {
        "F5" : lambda s: _f(s.fecha_alta),
        "F7" : lambda s: s.numero_contrato,
        "C9" : lambda s: e.razon_social,
        "E10": lambda s: e.giro,
        "C11": lambda s: e.nacionalidad,
        "C12": lambda s: e.fecha_constitucion,
        "C13": lambda s: e.rfc,
        "C19": lambda s: rl.nombre_completo,
        "C25": lambda s: e.telefono,
        "C26": lambda s: e.correo,
        # campos fijos
        "C30": lambda s: LLENAR_POR_ASESOR,  # Nombre persona de contacto
        "C31": lambda s: LLENAR_POR_ASESOR,  # Propietarios reales
        "C32": lambda s: LLENAR_POR_ASESOR,  # Administradores
        "C33": lambda s: LLENAR_POR_ASESOR,  # Director general
        "C34": lambda s: LLENAR_POR_ASESOR,  # Gerentes
        "C35": lambda s: LLENAR_POR_ASESOR,  # Persona de contacto
    }}

def _kyc_pm(sol):
    e = sol.empresa or Empresa()
    return {"Hoja1": {
        "O8" : lambda s: s.numero_contrato,
        "B11": lambda s: e.razon_social,
        "K11": lambda s: "JUAN JAVIER GILBERTO TELLEZ LOPEZ",
        "D13": lambda s: "MORAL",
        "E20": lambda s: e.giro,
    }}

def _perfil_pm(sol):
    e = sol.empresa or Empresa()
    rl = sol.representantes[0] if sol.representantes else RepresentanteLegal()
    return {"Hoja1": {
        "D146": lambda s: e.razon_social,
        "D147": lambda s: e.fecha_constitucion,
        "D148": lambda s: e.giro,
        "D150": lambda s: s.numero_contrato,
        "D151": lambda s: s.fecha_alta,
        "D152": lambda s: e.correo,
        "D156": lambda s: rl.nombre_completo,
    }}

MAPAS_XLSX = {
    "Checklist Expediente (3)": _checklist,      # PF: "1-Checklist Expediente (3).xlsx"
    "Checklist Expediente PM":  _checklist_pm,   # PM: renombrar a "1-Checklist Expediente PM.xlsx"
    "kyc Visita Ocular  (Firma)": _kyc,           # PF
    "CV kyc Visita Ocular":     _kyc_pm,          # PM
    "Perfil Persona Fisica":    _perfil,          # PF
    "CV Perfil Persona Moral":  _perfil_pm,       # PM
}


# ══════════════════════════════════════════════════════════════════════════════
# RENDERIZADO
# ══════════════════════════════════════════════════════════════════════════════

def render_docx(plantilla: Path, salida: Path, sol: Solicitud) -> Path:
    from docxtpl import DocxTemplate
    doc = DocxTemplate(str(plantilla))
    ctx = contexto_docx_pm(sol) if sol.tipo_persona == TipoPersona.MORAL else contexto_docx(sol)
    doc.render(ctx)
    salida.parent.mkdir(parents=True, exist_ok=True)
    doc.save(str(salida))
    return salida

def _mapa_para(nombre: str):
    for clave, fn in MAPAS_XLSX.items():
        if clave.lower() in nombre.lower():
            return fn
    return None

def render_xlsx(plantilla: Path, salida: Path, sol: Solicitud) -> Path:
    import openpyxl
    fn_mapa = _mapa_para(plantilla.name)
    salida.parent.mkdir(parents=True, exist_ok=True)
    es_macro = plantilla.suffix.lower() == ".xlsm"
    wb = openpyxl.load_workbook(str(plantilla), keep_vba=es_macro)
    if fn_mapa:
        for hoja, celdas in fn_mapa(sol).items():
            if hoja not in wb.sheetnames: continue
            ws = wb[hoja]
            for celda, getter in celdas.items():
                try:    valor = getter(sol)
                except: valor = ""
                if valor not in (None, ""): ws[celda] = valor
    wb.save(str(salida))
    return salida


# ══════════════════════════════════════════════════════════════════════════════
# PIPELINE → ZIP
# ══════════════════════════════════════════════════════════════════════════════

def procesar_a_zip(pdf_bytes: bytes, carpeta_pf: Path, carpeta_pm: Path) -> tuple[bytes, Solicitud, list, list]:
    """Extrae, rellena formatos y devuelve (zip_bytes, solicitud, advertencias, errores)."""
    sol = detectar_y_extraer(pdf_bytes)
    carpeta_plantillas = carpeta_pm if sol.tipo_persona == TipoPersona.MORAL else carpeta_pf
    advertencias = sol.advertencias()
    errores = []
    formatos = []

    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        salida = tmp_path / "formatos"
        salida.mkdir()

        for plantilla in sorted(carpeta_plantillas.iterdir()):
            out = salida / plantilla.name
            try:
                if plantilla.suffix.lower() == ".docx":
                    formatos.append(render_docx(plantilla, out, sol))
                elif plantilla.suffix.lower() in (".xlsx", ".xlsm"):
                    formatos.append(render_xlsx(plantilla, out, sol))
            except Exception as e:
                errores.append(f"{plantilla.name}: {e}")

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for f in formatos:
                zf.write(f, f.name)
        return buf.getvalue(), sol, advertencias, errores


# ══════════════════════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════════════════════

CARPETA_PF = Path(__file__).parent / "templates" / "pf"
CARPETA_PM = Path(__file__).parent / "templates" / "pm"

st.markdown("# 🏦 Apertura de Cuentas INVEX")
st.markdown("**Persona Física y Moral** — Llenado automático de formatos")
st.markdown('<hr class="divider">', unsafe_allow_html=True)

# ── verificar plantillas ───────────────────────────────────────────────────
if not CARPETA_PF.exists() or not any(CARPETA_PF.iterdir()):
    st.error("⚠️ No se encontraron plantillas en `templates/pf/`.")
    st.stop()

# ── subir PDF ─────────────────────────────────────────────────────────────
st.markdown("### 📄 Solicitud de Apertura")
st.markdown('<div class="upload-hint">Arrastra o selecciona el PDF de la Solicitud de Apertura (Persona Física o Moral)</div>',
            unsafe_allow_html=True)

archivo = st.file_uploader("", type=["pdf"], label_visibility="collapsed")

if archivo is None:
    st.stop()

pdf_bytes = archivo.read()

# ── procesar ──────────────────────────────────────────────────────────────
with st.spinner("Extrayendo datos del PDF..."):
    try:
        zip_bytes, sol, advertencias, errores = procesar_a_zip(pdf_bytes, CARPETA_PF, CARPETA_PM)
    except Exception as e:
        st.error(f"Error al procesar el PDF: {e}")
        st.stop()

es_pm = sol.tipo_persona == TipoPersona.MORAL
st.markdown('<hr class="divider">', unsafe_allow_html=True)

# ── datos extraídos ───────────────────────────────────────────────────────
tipo_label = "🏢 Persona Moral" if es_pm else "👤 Persona Física"
st.markdown(f"### {tipo_label} — Datos extraídos")

col1, col2 = st.columns(2)

if es_pm:
    e = sol.empresa or Empresa()
    rl = sol.representantes[0] if sol.representantes else RepresentanteLegal()
    with col1:
        st.markdown(f'<div class="campo-card"><div class="campo-label">Razón Social</div>'
                    f'<div class="campo-valor">{e.razon_social or "—"}</div></div>', unsafe_allow_html=True)
        st.markdown(f'<div class="campo-card"><div class="campo-label">RFC</div>'
                    f'<div class="campo-valor">{e.rfc or "—"}</div></div>', unsafe_allow_html=True)
        st.markdown(f'<div class="campo-card"><div class="campo-label">Fecha de constitución</div>'
                    f'<div class="campo-valor">{_f(e.fecha_constitucion) or "—"}</div></div>', unsafe_allow_html=True)
        st.markdown(f'<div class="campo-card"><div class="campo-label">Giro</div>'
                    f'<div class="campo-valor">{e.giro or "—"}</div></div>', unsafe_allow_html=True)
        st.markdown(f'<div class="campo-card"><div class="campo-label">Teléfono</div>'
                    f'<div class="campo-valor">{e.telefono or "—"}</div></div>', unsafe_allow_html=True)
    with col2:
        st.markdown(f'<div class="campo-card"><div class="campo-label">No. Contrato</div>'
                    f'<div class="campo-valor">{sol.numero_contrato or "—"}</div></div>', unsafe_allow_html=True)
        st.markdown(f'<div class="campo-card"><div class="campo-label">Fecha de alta</div>'
                    f'<div class="campo-valor">{_f(sol.fecha_alta) or "—"}</div></div>', unsafe_allow_html=True)
        st.markdown(f'<div class="campo-card"><div class="campo-label">Representante Legal</div>'
                    f'<div class="campo-valor">{rl.nombre_completo or "—"}</div></div>', unsafe_allow_html=True)
        st.markdown(f'<div class="campo-card"><div class="campo-label">Correo</div>'
                    f'<div class="campo-valor">{e.correo or "—"}</div></div>', unsafe_allow_html=True)
        st.markdown(f'<div class="campo-card"><div class="campo-label">Domicilio</div>'
                    f'<div class="campo-valor">{e.domicilio.una_linea() or "—"}</div></div>', unsafe_allow_html=True)
    if sol.accionistas:
        for ac in sol.accionistas:
            st.markdown(f'<div class="campo-card"><div class="campo-label">Accionista ({ac.porcentaje:.0f}%)</div>'
                        f'<div class="campo-valor">{ac.nombre_completo}</div></div>', unsafe_allow_html=True)
else:
    t = sol.titular
    with col1:
        st.markdown(f'<div class="campo-card"><div class="campo-label">Nombre completo</div>'
                    f'<div class="campo-valor">{t.nombre_completo or "—"}</div></div>', unsafe_allow_html=True)
        st.markdown(f'<div class="campo-card"><div class="campo-label">RFC</div>'
                    f'<div class="campo-valor">{t.rfc or "—"}</div></div>', unsafe_allow_html=True)
        st.markdown(f'<div class="campo-card"><div class="campo-label">CURP</div>'
                    f'<div class="campo-valor">{t.curp or "—"}</div></div>', unsafe_allow_html=True)
        st.markdown(f'<div class="campo-card"><div class="campo-label">Fecha de nacimiento</div>'
                    f'<div class="campo-valor">{_f(t.fecha_nacimiento) or "—"}</div></div>', unsafe_allow_html=True)
        st.markdown(f'<div class="campo-card"><div class="campo-label">Teléfono</div>'
                    f'<div class="campo-valor">{t.telefono or "—"}</div></div>', unsafe_allow_html=True)
    with col2:
        st.markdown(f'<div class="campo-card"><div class="campo-label">No. Contrato</div>'
                    f'<div class="campo-valor">{sol.numero_contrato or "—"}</div></div>', unsafe_allow_html=True)
        st.markdown(f'<div class="campo-card"><div class="campo-label">Fecha de alta</div>'
                    f'<div class="campo-valor">{_f(sol.fecha_alta) or "—"}</div></div>', unsafe_allow_html=True)
        st.markdown(f'<div class="campo-card"><div class="campo-label">INE / Identificación</div>'
                    f'<div class="campo-valor">{t.numero_identificacion or "—"}</div></div>', unsafe_allow_html=True)
        st.markdown(f'<div class="campo-card"><div class="campo-label">Correo</div>'
                    f'<div class="campo-valor">{t.correo or "—"}</div></div>', unsafe_allow_html=True)
        st.markdown(f'<div class="campo-card"><div class="campo-label">Domicilio</div>'
                    f'<div class="campo-valor">{t.domicilio.una_linea() or "—"}</div></div>', unsafe_allow_html=True)
    if sol.cotitulares:
        st.markdown(f'<div class="campo-card"><div class="campo-label">Cotitular</div>'
                    f'<div class="campo-valor">{sol.cotitulares[0].nombre_completo}</div></div>',
                    unsafe_allow_html=True)

# ── advertencias / errores ────────────────────────────────────────────────
if advertencias:
    for w in advertencias:
        st.markdown(f'<div class="warn-box">⚠️ {w}</div>', unsafe_allow_html=True)

if errores:
    for e in errores:
        st.markdown(f'<div class="warn-box">❌ {e}</div>', unsafe_allow_html=True)

# ── formatos generados ────────────────────────────────────────────────────
st.markdown('<hr class="divider">', unsafe_allow_html=True)
carpeta_activa = CARPETA_PM if es_pm else CARPETA_PF
n_formatos = len([f for f in carpeta_activa.iterdir() if f.suffix.lower() in (".docx",".xlsx",".xlsm")])
n_ok = n_formatos - len(errores)

if n_ok > 0:
    st.markdown(f'<div class="ok-box">✅ {n_ok} de {n_formatos} formatos generados correctamente.</div>',
                unsafe_allow_html=True)

# ── descarga ──────────────────────────────────────────────────────────────
st.markdown("### 📦 Descarga")
nombre_zip = f"EXPEDIENTE_{sol.numero_contrato or 'SC'}.zip"

st.download_button(
    label=f"⬇️  Descargar {nombre_zip}",
    data=zip_bytes,
    file_name=nombre_zip,
    mime="application/zip",
    use_container_width=True,
)
